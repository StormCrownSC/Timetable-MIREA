#!/usr/bin/env python3
import os
import re
import time
import json
from datetime import datetime
import requests
from bs4 import BeautifulSoup
import openpyxl
from concurrent.futures import ThreadPoolExecutor as Pool
import psycopg2
from psycopg2 import Error
import pyexcel as p


class Parser:
    def __init__(self):
        self.log("start parser")
        self.links = dict()
        self.timetable = dict()
        self.lecturer_list = set()
        self.lessons_list = self.list_of_lessons()
        self.degree_list = ["Бакалавриат", "Специалитет", "Магистратура", "Аспирантура"]
        self.course = 0

        self.main()

    def main(self):
        self.parse_site()
        self.loader()
        self.parse_data()
        self.open_database()

    @staticmethod
    def log(text):
        print(datetime.utcnow().isoformat(sep=' ', timespec='milliseconds') + " UTC LOG: " + text)

    def parse_site(self):
        self.log("parse site")
        page = BeautifulSoup(requests.get("https://www.mirea.ru/schedule/").text, "html.parser")
        data_of_institute = page.findAll('a', attrs={"class": "uk-text-bold"})
        for elem in data_of_institute:
            temp_data = re.findall(r"Институт [\w \.]+", str(elem))
            if len(temp_data) > 0:
                temp_data_of_link = elem.find_parent("div").find_parent("div").find_all_next('a', attrs={"class": "uk-link-toggle"})
                temp_link = re.findall(r'<a class="uk-link-toggle" href="([\S]+.xl[\w]+)" target="_blank">', str(temp_data_of_link))
                if len(temp_link) > 0:
                    for i, item in enumerate(temp_data_of_link):
                        tmp = item.find('div', attrs={"class": "uk-link-heading uk-margin-small-top"})
                        temp_curs = re.findall(r"(\d) курс", str(tmp) if tmp is not None else "")
                        if len(temp_curs) > 0:
                            self.links[temp_link[i]] = {"institute": temp_data[0], "curs": temp_curs[0]}

    @staticmethod
    def download(link):
        link, data = link
        with open("temp/" + data["institute"] + "_-_" + data["curs"] + "_-_" + link.split("/")[-1], "wb") as file:
            link_content = requests.get(link)
            count_request = 0
            while link_content.status_code != 200 and count_request != 100:
                link_content = requests.get(link)
                count_request += 1
            file.write(link_content.content)


    def loader(self):
        self.log("start loader")
        self.check_directory()
        with Pool(max_workers=len(self.links)) as executor:
            executor.map(self.download, self.links.items())

    @staticmethod
    def remove_file(file):
        os.remove("temp/" + file)

    def check_directory(self):
        if os.path.exists("success"):
            os.remove("success")
        if not os.path.exists("temp"):
            os.mkdir("temp")
        elif len(os.listdir("temp")) != 0:
            with Pool(max_workers=len(os.listdir("temp"))) as executor:
                executor.map(self.remove_file, os.listdir("temp"))

    def data_of_lesson(self, sheet, item, col, global_col):
        try:
            new_item = item
            if sheet.cell(row=new_item, column=global_col).value is None:
                new_item -= 1
            lesson_number = sheet.cell(row=new_item, column=global_col).value
            if sheet.cell(row=new_item-2, column=col).value == 7 or lesson_number == 8:
                lesson_number += 1
            tmp_of_lecturer = sheet.cell(row=item, column=col+2).value
            lecturer_temp = re.findall(r"[\w]+. \w.\w.", tmp_of_lecturer) if tmp_of_lecturer is not None else []
            
            for tmp in lecturer_temp:
                self.lecturer_list.add(tmp.rstrip())
            self.lecturer_list = set(lecturer_list)

            lesson_title = ' '.join(str(sheet.cell(row=item, column=col).value).split())
            type_of_lesson = ' '.join(str(sheet.cell(row=item, column=col+1).value).split())
            auditorium = ' '.join(str(sheet.cell(row=item, column=col+3).value).split())
            
            temp_array = [lesson_title, type_of_lesson, ' '.join(lecturer_temp), auditorium]
            type_of_week = 2
            if sheet.cell(row=item, column=global_col+3).value == "I":
                type_of_week = 1
            return type_of_week, lesson_number, temp_array
        except:
            return None, None, None
        
    def read_files(self, file):
        try:
            if file[-1:] != "x":
                p.save_book_as(file_name="temp/" + file, dest_file_name="temp/" + file + "x")
                os.remove("temp/" + file)
                file += "x"
            book = openpyxl.load_workbook("temp/" + file)
            institute_name, course_num, temple = file.split("_-_")
            self.course = int(course_num) if int(course_num) > self.course else self.course
            for name in book.sheetnames:
                sheet = book[name]
                flag = False

                for index, arr in enumerate(sheet):
                    flag_optimize = False
                    global_col = None
                    max_row = sheet.max_row
                    first_len_flag = True
                    for elem in arr:
                        if len(re.findall(r"\w\w\w\w\-\d\d\-\d\d", str(elem.value))) != 0:
                            flag = True
                            flag_optimize = True
                            col = elem.column
                            letter = elem.row
                            if global_col is None:
                                global_col = elem.column - 4
                            temp_dict = {1: {}, 2: {}}
                            day_index = 0
                            for item in range(letter + 2, max_row):
                                if first_len_flag is True and sheet.cell(row=item, column=global_col).value is None \
                                    and sheet.cell(row=item-1, column=global_col).value is None:
                                    first_len_flag = False
                                    max_row = item
                                    break
                                if sheet.cell(row=item, column=global_col).value is not None and \
                                    re.search(r'1', str(sheet.cell(row=item, column=global_col).value)):
                                    day_index += 1
                                    temp_dict[1][day_index] = {}
                                    temp_dict[2][day_index] = {}
                                type_of_week, lesson_number, tmp_data = self.data_of_lesson(sheet, item, col, global_col)
                                if type(lesson_number) == int:
                                    temp_dict[type_of_week][day_index][lesson_number] = tmp_data
                            
                            enter_data = {"data": temp_dict, "institute": institute_name, "course": course_num}
                            self.timetable[re.findall(r"\w\w\w\w\-\d\d\-\d\d", str(elem.value))[0]] = enter_data
                        if flag_optimize is False and index > 10:
                            break
                    if flag is True:
                        break
            return 1
        except:
            return 0

    def parse_data(self):
        self.log("start parse data")
        list_of_file = os.listdir("temp")
        bad_files = 0
        for index, elem in enumerate(list_of_file):
            if self.read_files(elem) == 0:
                bad_files += 1
        self.log("read " + str(round(((len(list_of_file) - bad_files) / len(list_of_file)) * 100)) + "% of files")
    
    @staticmethod
    def degree_of_letter(letter):
        if letter == "Б":
            return "Бакалавриат"
        if letter == "С":
            return "Специалитет"
        if letter == "М":
            return "Магистратура"
        return "Аспирантура"

    def timetable_table(self, connection):
        with connection.cursor() as cursor:
            for num, arr in self.timetable.items():
                edit = False
                insert_query = "INSERT INTO timetable (id_to_group, subject_to_number, id_lectur, subject_title, auditorium, day_week, type_of_week) VALUES"
                cursor.execute("SELECT id_group FROM study_group WHERE group_name = '" + str(num) + "'")
                id_group = str(cursor.fetchone())[1:-2]
                for type_of_week, type_of_week_arr in arr["data"].items():
                    for day_week, day_array in type_of_week_arr.items():
                        for subject_to_number, subject_number_array in day_array.items():
                            if cursor.execute("SELECT COUNT(*) FROM timetable WHERE id_to_group = '" + str(id_group) + "' and " + \
                                "subject_to_number = '" + str(subject_to_number) + "' and " + "day_week = '" + str(day_week) + "' and " + \
                                "type_of_week = '" + str(type_of_week) + "'") is None:
                                edit = True
                                subject_title = subject_number_array[0]
                                auditorium = subject_number_array[3]
                                cursor.execute("SELECT id_lecturer FROM lecturer \
                                    WHERE full_name = '" + str(subject_number_array[2]) + "'")
                                id_lectur = str(cursor.fetchone())[1:-2]
                                if not id_lectur.isdigit():
                                    id_lectur = "NULL"
                                else:
                                    id_lectur = "'" + id_lectur + "'"

                                insert_query += " ('" + str(id_group) + "', '" + str(subject_to_number) + "', " + str(id_lectur) + \
                                    ", '" + str(subject_title) + "', '" + str(auditorium) + "', '" + str(day_week) + "', '" + \
                                        str(type_of_week) + "'),"
                if edit is True:
                    insert_query = insert_query[:-1]
                    cursor.execute(insert_query)
                    connection.commit()
            self.log("Таблица пар успешно заполнена")

    def study_group_table(self, connection):
        with connection.cursor() as cursor:
            for num, elem in self.timetable.items():
                if cursor.execute("SELECT COUNT(*) FROM study_group WHERE group_name = '" + str(num) + "'") is None:
                    insert_query = "INSERT INTO study_group (group_name, id_institute, id_of_course, id_degree) VALUES"
                    cursor.execute("SELECT id_of_the_institute FROM institute \
                        WHERE name_of_the_institute = '" + str(elem["institute"]) + "'")
                    id_institute = str(cursor.fetchone())[1:-2]
                    id_institute = id_institute if id_institute.isdigit() else "1"
                    id_of_course = elem["course"]
                    cursor.execute("SELECT id_of_degree FROM degree \
                        WHERE degree_of_study = '" + str(self.degree_of_letter(num[2])) + "'")
                    id_degree = str(cursor.fetchone())[1:-2]
                    insert_query += " ('" + str(num) + "', '" + str(id_institute) + "', '" + str(id_of_course) + "', '" + str(id_degree) + "'),"
                    insert_query = insert_query[:-1]
                    cursor.execute(insert_query)
                    connection.commit()
            self.log("Таблица групп успешно заполнена")

    def institute_table(self, connection):
        with connection.cursor() as cursor:
            for elem in self.links.values():
                if cursor.execute("SELECT COUNT(*) FROM institute WHERE name_of_the_institute = '" + str(elem["institute"]) + "'") is None:
                    insert_query = "INSERT INTO institute (name_of_the_institute) VALUES"
                    insert_query += " ('" + str(elem["institute"]) + "'),"
                    insert_query = insert_query[:-1]
                    cursor.execute(insert_query)
                    connection.commit()
            self.log("Таблица институтов успешно заполнена")

    def course_table(self, connection):
        with connection.cursor() as cursor:
            for elem in range(1, self.course + 1):
                if cursor.execute("SELECT COUNT(*) FROM course WHERE id_of_course = '" + str(elem) + "'") is None:
                    insert_query = "INSERT INTO course (id_of_course) VALUES"
                    insert_query += " ('" + str(elem) + "'),"
                    insert_query = insert_query[:-1]
                    cursor.execute(insert_query)
                    connection.commit()
            self.log("Таблица курсов успешно заполнена")

    def degree_table(self, connection):
        with connection.cursor() as cursor:
            for elem in self.degree_list:
                if cursor.execute("SELECT COUNT(*) FROM degree WHERE degree_of_study = '" + str(elem) + "'") is None:
                    insert_query = "INSERT INTO degree (degree_of_study) VALUES"
                    insert_query += " ('" + str(elem) + "'),"
                    insert_query = insert_query[:-1]
                    cursor.execute(insert_query)
                    connection.commit()
            self.log("Таблица уровней обучения успешно заполнена")

    def call_schedule_table(self, connection):
        with connection.cursor() as cursor:
            for num, elem in self.lessons_list.items():
                if cursor.execute("SELECT COUNT(*) FROM call_schedule WHERE subject_number = '" + str(num) + "'") is None:
                    insert_query = "INSERT INTO call_schedule (subject_number, time_start, time_end) VALUES"
                    insert_query += " ('" + str(num) + "', '" + str(elem[0]) + "', '" + str(elem[1]) + "'),"
                    insert_query = insert_query[:-1]
                    cursor.execute(insert_query)
                    connection.commit()
            self.log("Таблица звонков успешно заполнена")

    def lecturer_table(self, connection):
        with connection.cursor() as cursor:
            for elem in self.lecturer_list:
                if cursor.execute("SELECT COUNT(*) FROM lecturer WHERE full_name = '" + str(elem) + "'") is None:
                    insert_query = "INSERT INTO lecturer (full_name) VALUES"
                    insert_query += " ('" + str(elem) + "'),"
                    insert_query = insert_query[:-1]
                    cursor.execute(insert_query)
                    connection.commit()
            self.log("Таблица преподавателей успешно заполнена")

    def open_database(self):
        try:
            self.log("Подключение к базе данных")
            with psycopg2.connect(user="denis",
                                        # пароль, который указали при установке PostgreSQL
                                        password="password",
                                        host="database",
                                        port="5432",
                                        database="TimeTableDB") as connection:

                self.lecturer_table(connection)
                self.call_schedule_table(connection)
                self.degree_table(connection)
                self.institute_table(connection)
                self.course_table(connection)
                self.study_group_table(connection)
                self.timetable_table(connection)
                self.log("Соединение с PostgreSQL закрыто")

        except (Exception, Error) as error:
            self.log("Ошибка при работе с PostgreSQL")

    @staticmethod
    def list_of_lessons():
        return {1: ["9:00", "10:30"], 2: ["10:40", "12:10"], 3: ["12:40", "14:10"], 4: ["14:20", "15:50"], 
        5: ["16:20", "17:50"], 6: ["18:00", "19:30"], 7: ["19:40", "21:10"], 8: ["18:30", "20:00"], 9: ["20:10", "21:40"]}

    def __del__(self):
        self.check_directory()
        if os.path.exists("temp"):
            os.rmdir("temp")
        with open("success", "w") as file:
            file.write("")
        self.log("end")


if __name__ == "__main__":
    try:
        while True:
            Parser()
            time.sleep(14400)
    except KeyboardInterrupt():
        pass